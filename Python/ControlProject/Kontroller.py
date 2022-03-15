from openpyxl import load_workbook
from os import getcwd, walk
from shutil import copyfile
from datetime import date, datetime
from mailAPI import send_message, service

### Variables ###

sheet = load_workbook(filename="Kontroller.xlsx")
ws = sheet.active

today = date.today()
todayDate = int(today.strftime("%d%m%Y"))

mailingList = []
contactInfo = {}
controlsInSheet = []

filesToSend = []
maxControlRow = len(ws['A'])
maxContactsRow = len(ws['O'])


### Classes ###


class Ctrls:
    def __init__(self, number, control, responsible, due, verification):
        self.number = number
        self.control = control
        self.responsible = responsible
        self.due = due
        self.verification = verification


### Functions ###

def createCtrls():
    for value in ws.iter_rows(min_row=2,
                              max_row=maxControlRow,
                              min_col=1,
                              max_col=5,
                              values_only=True):
        classValue = Ctrls(value[0],
                           value[1],
                           value[2],
                           value[3],
                           value[4])

        controlsInSheet.append(classValue)


def checkForDue(value0, value1, value2, value3, value4):
    global status
    if value4 != "X":
        dueDateStr = datetime.strptime(value3, '%d.%m.%Y')
        dueDateint = int(dueDateStr.strftime("%d%m%Y"))
        if dueDateint - todayDate == 0:
            status = "Send The email!"
            if value2 in contactInfo:
                mailingList.append([value0, value1, value2, value3, contactInfo[value2]])
            else:
                print("Update needed for {}".format(value2))
        elif dueDateint - todayDate == 10000000:
            status = "Send a reminder! He got 10 days left"
            if value2 in contactInfo:
                mailingList.append([value0, value1, value2, value3, contactInfo[value2]])
            else:
                print("Update needed for {}".format(value2))
        elif dueDateint - todayDate == 5000000:
            status = "Send a reminder!"
        else:
            status = "Take it easy"

def contactInfoFunc():
    for value, val in ws.iter_rows(
            min_row=2,
            max_row=maxContactsRow,
            min_col=15,
            max_col=16,
            values_only=True):
        if value == None:
            pass

        else:
            contactInfo[value] = val


def makeControlDoc():
    for item in mailingList:
        control = item[1] + ".xlsx"
        controlTitle = str(item[0]) + " " + item[1] + " " + item[3] + ".xlsx"
        for roots, dirs, files in walk("."):
            for file in files:
                if control in file:
                    original = getcwd() + "\Templates" + "\\" + file
                    target = getcwd() + "\Temps" + "\\" + controlTitle
                    copyfile(original, target)
                    filesToSend.append(target)




### Logic ###
contactInfoFunc()
createCtrls()
for i in range(maxControlRow - 1):
    checkForDue(controlsInSheet[i].number,
                controlsInSheet[i].control,
                controlsInSheet[i].responsible,
                controlsInSheet[i].due,
                controlsInSheet[i].verification)

makeControlDoc()

for i in filesToSend:
    send_message(service, "chr.maints@gmail.com", i.split("\\")[5],
                 status, [i])